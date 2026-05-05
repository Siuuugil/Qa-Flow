import click
from rich.console import Console
from dotenv import load_dotenv
from pathlib import Path
import os

console = Console()


@click.command()
@click.option("--mode",
              type=click.Choice(["ai-only", "full"]),
              default="ai-only",
              help="실행 모드 선택")
@click.option("--report",
              is_flag=True,
              default=False,
              help="리포트 파일 생성")
@click.option("--provider",
              type=click.Choice(["claude", "gemini"]),
              default=None,
              help="AI Provider 선택 (기본값: .env 설정 따름)")
@click.option("--focus",
              type=click.Choice(["structure",
                                 "convention",
                                 "security",
                                 "bug",
                                 "qa",
                                 "dev",
                                 "performance",
                                 "test",
                                 "general"]),
              default=None,
              help="리뷰 관점 선택")
@click.option("--file",
              default=None,
              help="특정 파일 직접 분석 (git diff 없이)")
@click.option("--tc",
              is_flag=True,
              default=False,
              help="분석 결과 기반 TC 엑셀 자동 생성")
@click.option("--jira",
              is_flag=True,
              default=False,
              help="생성된 TC를 Jira 이슈로 자동 등록")
def scan(mode, report, provider, focus, file, tc, jira):
    """코드 분석 실행"""

    load_dotenv(dotenv_path=Path.cwd() / ".env", encoding="utf-8")

    ai_provider = provider or os.getenv("AI_PROVIDER", "gemini")

    if file:
        console.print("\n[bold cyan]파일 분석 모드[/bold cyan]")
        console.print(f"파일: [yellow]{file}[/yellow]")
        console.print(f"AI Provider: [yellow]{ai_provider}[/yellow]")
        console.print(f"Focus: [yellow]{focus or 'default'}[/yellow]\n")

        convention_result = ""
        ai_result = ""

        if mode == "full":
            convention_result = run_convention_check_file(file)
            ai_result = run_ai_review_file(ai_provider, focus, file)
            if report:
                run_report(ai_result, convention_result, focus, mode, auto_tc=tc, provider=ai_provider)
        else:
            ai_result = run_ai_review_file(ai_provider, focus, file)
            if tc:
                run_tc_generation(ai_provider, ai_result, "", jira)

    else:
        console.print("\n[bold cyan]QA-Flow 분석 시작[/bold cyan]")
        console.print(f"모드: [yellow]{mode}[/yellow]")
        console.print(f"AI Provider: [yellow]{ai_provider}[/yellow]")
        console.print(f"Focus: [yellow]{focus or 'default'}[/yellow]")
        console.print(f"리포트 생성: [yellow]{report}[/yellow]\n")

        convention_result = ""
        ai_result = ""

        if mode == "ai-only":
            ai_result = run_ai_review(ai_provider, focus)
            if tc:
                run_tc_generation(ai_provider, ai_result, "", jira)
        elif mode == "full":
            convention_result = run_convention_check()
            ai_result = run_ai_review(ai_provider, focus)
            if report:
                run_report(ai_result, convention_result, focus, mode, auto_tc=tc, provider=ai_provider)


def run_ai_review(provider: str, focus: str = None):
    """AI 코드 리뷰 실행"""
    console.print("[bold]AI 코드 리뷰 실행 중...[/bold]")

    from cli.core.ai_review import AIReview

    with console.status("[bold green]AI가 코드를 분석하고 있습니다..."):
        reviewer = AIReview(provider=provider, focus=focus)
        result = reviewer.review()

    console.print(result)
    return result


def run_ai_review_file(provider: str, focus: str = None, file: str = None):
    """특정 파일 AI 분석"""

    allowed_extensions = (
        ".py", ".java", ".js", ".ts", ".jsx", ".tsx",
        ".go", ".rs", ".kt", ".swift", ".cpp", ".c", ".cs"
    )
    if not file.endswith(allowed_extensions):
        console.print(f"[red]지원하지 않는 파일 형식입니다: {file}[/red]")
        return ""

    try:
        with open(file, "r", encoding="utf-8") as f:
            code = f.read()
    except FileNotFoundError:
        console.print(f"[red]파일을 찾을 수 없습니다: {file}[/red]")
        return ""
    except Exception as e:
        console.print(f"[red]파일 읽기 오류: {str(e)}[/red]")
        return ""

    from cli.core.ai_review import AIReview

    with console.status(f"[bold green]AI가 {file} 을 분석하고 있습니다..."):
        reviewer = AIReview(provider=provider, focus=focus)
        result = reviewer.provider.review(code)

    console.print(result)
    return result


def run_convention_check():
    """컨벤션 체크 실행"""
    console.print("[bold]컨벤션 체크 실행 중...[/bold]")

    from cli.core.convention import ConventionChecker

    checker = ConventionChecker()
    result = checker.check()

    console.print(result)
    return result


def run_convention_check_file(file: str):
    """특정 파일 컨벤션 체크"""
    console.print("[bold]컨벤션 체크 실행 중...[/bold]")

    import subprocess

    result = ""

    if file.endswith((".js", ".ts", ".jsx", ".tsx")):
        try:
            r = subprocess.run(
                ["npx", "eslint", file],
                capture_output=True, text=True
            )
            if r.returncode == 0:
                result = "[ESLint] 컨벤션 위반 없음"
            else:
                result = f"[ESLint] 위반 사항:\n{r.stdout}"
        except FileNotFoundError:
            result = "[ESLint] 설치되지 않아 스킵합니다."

    elif file.endswith(".py"):
        try:
            r = subprocess.run(
                ["flake8", "--max-line-length=100", file],
                capture_output=True, text=True
            )
            if r.returncode == 0:
                result = "[Flake8] 컨벤션 위반 없음"
            else:
                result = f"[Flake8] 위반 사항:\n{r.stdout}"
        except FileNotFoundError:
            result = "[Flake8] 설치되지 않아 스킵합니다."

    elif file.endswith(".java"):
        result = "[Java] 컨벤션 체크는 AI 리뷰로 대체합니다."

    console.print(result)
    return result


def run_report(ai_result: str = "", convention_result: str = "", focus: str = None, mode: str = "full", auto_tc: bool = False, provider: str = None):
    """리포트 생성"""
    console.print("[bold]리포트 생성 중...[/bold]")

    from cli.core.reporter import Reporter

    reporter = Reporter()
    report_path = reporter.generate(
        ai_result=ai_result,
        convention_result=convention_result,
        focus=focus or "general",
        mode=mode
    )

    console.print("[green]리포트 생성 완료![/green]")

    # --tc 없으면 물어보기
    if not auto_tc:
        import click
        if click.confirm("\n현재 리뷰 기준으로 TC를 만들어드릴까요?", default=False):
            auto_tc = True

    if auto_tc:
        if not ai_result or "변경된 코드가 없습니다" in ai_result:
            console.print("[yellow]TC 생성 건너뜀: AI 분석 결과가 없습니다.[/yellow]")
            return
        tc_path = run_tc_generation(provider or os.getenv("AI_PROVIDER", "gemini"), ai_result, convention_result)
        if report_path and tc_path:
            console.print(f"\n[dim]리포트: {report_path} → TC: {tc_path}[/dim]")


def run_tc_generation(provider: str, ai_result: str, convention_result: str = "", upload_jira: bool = False):
    """TC 엑셀 생성 및 Jira 등록"""
    from cli.core.tc_generator import TCGenerator, upload_to_jira

    if provider == "gemini":
        from cli.core.providers.gemini import GeminiProvider
        p = GeminiProvider(system_prompt="")
    else:
        from cli.core.providers.claude import ClaudeProvider
        p = ClaudeProvider(system_prompt="")

    console.print("\n[bold cyan]TC 자동 생성 시작[/bold cyan]")

    generator = TCGenerator(provider_instance=p)
    tc_list = generator.generate_from_analysis(ai_result, convention_result)

    if not tc_list:
        console.print("[red]TC 생성 실패: AI 응답을 파싱할 수 없습니다.[/red]")
        return

    console.print(f"[green]TC {len(tc_list)}개 생성 완료![/green]")

    excel_path = generator.export_excel(tc_list)
    if excel_path:
        console.print(f"[green]엑셀 저장 완료: {excel_path}[/green]")
    else:
        console.print("[red]엑셀 저장 실패[/red]")
        return

    if upload_jira:
        console.print("\n[bold cyan]Jira 이슈 등록 시작[/bold cyan]")
        result = upload_to_jira(tc_list, excel_path)
        created = result.get("created", {})
        failed = result.get("failed", [])

        if created:
            console.print(f"[green]Jira 등록 완료: {len(created)}건[/green]")
            for tc_id, issue_key in created.items():
                console.print(f"  [dim]{tc_id} → {issue_key}[/dim]")

        if failed:
            console.print(f"[yellow]등록 실패: {len(failed)}건 → {', '.join(failed)}[/yellow]")

        if created and excel_path:
            _update_excel_jira_keys(excel_path, created)

    console.print(f"\n[bold green]TC 작업 완료![/bold green] 파일: {excel_path}")
    return excel_path


def _update_excel_jira_keys(excel_path: str, created: dict):
    """엑셀 Jira_Issue 컬럼에 이슈 키 업데이트"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(excel_path)
        ws = wb.active
        # TC_ID는 1번 컬럼, Jira_Issue는 13번 컬럼
        for row in ws.iter_rows(min_row=2):
            tc_id = row[0].value
            if tc_id and tc_id in created:
                row[12].value = created[tc_id]
        wb.save(excel_path)
    except Exception as e:
        console.print(f"[dim]엑셀 Jira 키 업데이트 실패: {str(e)}[/dim]")
