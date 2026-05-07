import click
import json
import os
from rich.console import Console
from pathlib import Path
from dotenv import load_dotenv

console = Console()


@click.command()
def jira():
    """마지막 TC 기반 Jira 이슈 등록"""

    load_dotenv(dotenv_path=Path.cwd() / ".env", encoding="utf-8")

    # 가장 최근 TC 파일 찾기
    tc_dir = Path("qa-reports")
    if not tc_dir.exists():
        console.print("[red]TC 파일이 없습니다. 먼저 qa-flow tc 를 실행하세요.[/red]")
        return

    tc_files = sorted(tc_dir.glob("*.xlsx"), key=os.path.getmtime, reverse=True)
    if not tc_files:
        console.print("[red]TC 파일이 없습니다. 먼저 qa-flow tc 를 실행하세요.[/red]")
        return

    latest = tc_files[0]

    # 확인 질문
    if not click.confirm(f"\n마지막 TC: [bold]{latest.name}[/bold] 기준으로 Jira에 등록하시겠습니까?", default=False):
        console.print("[dim]취소했습니다.[/dim]")
        return

    # 엑셀에서 TC 읽기
    try:
        from openpyxl import load_workbook
        wb = load_workbook(latest)
        ws = wb.active

        tc_list = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            tc_list.append({
                "test_scenario": row[2] or "",
                "module": row[1] or "",
                "precondition": row[3] or "",
                "test_steps": row[4] or "",
                "test_data": row[5] or "",
                "expected_result": row[6] or "",
                "priority": row[8] or "Major",
                "severity": row[9] or "Major",
                "test_type": row[10] or "기능",
            })
    except Exception as e:
        console.print(f"[red]TC 파일 읽기 실패: {str(e)}[/red]")
        return

    if not tc_list:
        console.print("[red]TC 데이터가 없습니다.[/red]")
        return

    console.print(f"[green]TC {len(tc_list)}개 로드 완료![/green]")

    from cli.core.tc_generator import upload_to_jira
    from cli.commands.scan import _update_excel_jira_keys
    result = upload_to_jira(tc_list, str(latest))
    created = result.get("created", {})
    failed = result.get("failed", [])

    if created:
        console.print(f"[green]Jira 등록 완료: {len(created)}건[/green]")
        for tc_id, issue_key in created.items():
            console.print(f"  [dim]{tc_id} → {issue_key}[/dim]")

    if failed:
        console.print(f"[yellow]등록 실패: {len(failed)}건[/yellow]")