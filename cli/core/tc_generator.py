import os
import json
import re
from datetime import datetime
from rich.console import Console

console = Console()

TC_PROMPT = """당신은 QA 엔지니어입니다.
아래 코드 분석 결과를 바탕으로 실무 수준의 테스트 케이스를 생성하세요.

반드시 아래 JSON 형식으로만 답변하세요. 다른 텍스트는 절대 추가하지 마세요.

[
  {
    "module": "모듈명 (예: 로그인, 결제, 회원가입)",
    "test_scenario": "테스트 시나리오 (예: 정상 로그인 확인)",
    "precondition": "사전 조건 (예: 회원가입 완료된 계정 존재)",
    "test_steps": "1. 첫번째 단계\\n2. 두번째 단계\\n3. 세번째 단계",
    "test_data": "테스트 데이터 (예: ID: test@test.com, PW: Test1234!)",
    "expected_result": "예상 결과 (예: 메인 페이지로 이동, 환영 메시지 표시)",
    "priority": "Blocker 또는 Critical 또는 Major 또는 Minor 또는 Low",
    "severity": "Blocker 또는 Critical 또는 Major 또는 Minor 또는 Low",
    "test_type": "기능 또는 보안 또는 성능 또는 UI 또는 통합"
  }
]

Priority/Severity 기준 (Jira 표준):
- Blocker: 테스트 진행 불가, 시스템 크래시, 핵심 기능 완전 불가
- Critical: 핵심 기능 심각한 오작동, 데이터 손실 위험
- Major: 주요 기능 오작동, 우회 방법 없음
- Minor: 부분 기능 오작동, 우회 방법 존재
- Low: UI 오류, 사소한 불편, 기능에 영향 없음

코드 분석 결과:
{analysis}

위 분석을 바탕으로 최소 5개 이상의 테스트 케이스를 생성하세요.
JSON 배열만 반환하세요."""


class TCGenerator:
    def __init__(self, provider_instance):
        self.provider = provider_instance

    def generate_from_analysis(self, ai_result: str, convention_result: str = "") -> list:
        """AI 분석 결과로부터 TC 생성"""
        analysis = f"AI 리뷰:\n{ai_result}\n\n컨벤션 체크:\n{convention_result}"
        prompt = TC_PROMPT.format(analysis=analysis)

        with console.status("[bold green]AI가 테스트 케이스를 생성하고 있습니다..."):
            raw = self.provider.review(prompt)

        return self._parse_tc(raw)

    def _parse_tc(self, raw: str) -> list:
        """AI 응답에서 JSON 파싱"""
        try:
            # 마크다운 코드블록 제거
            cleaned = re.sub(r"```(?:json)?", "", raw).strip().rstrip("`").strip()
            # JSON 배열 추출
            match = re.search(r"\[.*\]", cleaned, re.DOTALL)
            if match:
                return json.loads(match.group())
        except Exception as e:
            console.print(f"[yellow]TC 파싱 경고: {str(e)}[/yellow]")
        return []

    def export_excel(self, tc_list: list, output_path: str = None) -> str:
        """TC 리스트를 실무 포맷 엑셀로 저장"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import (
                Font, PatternFill, Alignment, Border, Side
            )
            from openpyxl.utils import get_column_letter
        except ImportError:
            console.print("[red]openpyxl 설치 필요: pip install openpyxl[/red]")
            return ""

        if not output_path:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            os.makedirs("qa-reports", exist_ok=True)
            output_path = f"qa-reports/TC_{ts}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Test Cases"

        # 스타일 정의
        header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill("solid", start_color="1F4E79")
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        priority_colors = {
            "Blocker":  "FF0000",
            "Critical": "FF6600",
            "Major":    "FFA500",
            "Minor":    "FFD700",
            "Low":      "90EE90",
        }
        status_fill = PatternFill("solid", start_color="F2F2F2")

        # 헤더 정의
        headers = [
            ("TC_ID",           12),
            ("Module",          15),
            ("Test_Scenario",   30),
            ("Precondition",    25),
            ("Test_Steps",      40),
            ("Test_Data",       25),
            ("Expected_Result", 30),
            ("Actual_Result",   30),
            ("Priority",        12),
            ("Severity",        12),
            ("Test_Type",       12),
            ("Status",          14),
            ("Jira_Issue",      14),
            ("Remarks",         20),
        ]

        # 헤더 행 작성
        for col_idx, (col_name, col_width) in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
            ws.column_dimensions[get_column_letter(col_idx)].width = col_width

        ws.row_dimensions[1].height = 22

        # 데이터 행 작성
        for row_idx, tc in enumerate(tc_list, start=2):
            tc_id = f"TC-{row_idx - 1:03d}"
            priority = tc.get("priority", "Major")
            severity = tc.get("severity", "Major")
            status = "Not Executed"

            row_data = [
                tc_id,
                tc.get("module", ""),
                tc.get("test_scenario", ""),
                tc.get("precondition", ""),
                tc.get("test_steps", ""),
                tc.get("test_data", ""),
                tc.get("expected_result", ""),
                "",          # Actual_Result (수행 후 작성)
                priority,
                severity,
                tc.get("test_type", "기능"),
                status,
                "",          # Jira_Issue (연동 후 채워짐)
                "",          # Remarks
            ]

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                cell.alignment = left

                # Priority 컬럼 색상
                if col_idx == 9:
                    color = priority_colors.get(priority, "FFFFFF")
                    cell.fill = PatternFill("solid", start_color=color)
                    cell.font = Font(name="Arial", bold=True, size=9,
                                     color="FFFFFF" if priority in ("Blocker", "Critical") else "000000")
                    cell.alignment = center

                # Severity 컬럼 색상
                elif col_idx == 10:
                    color = priority_colors.get(severity, "FFFFFF")
                    cell.fill = PatternFill("solid", start_color=color)
                    cell.font = Font(name="Arial", bold=True, size=9,
                                     color="FFFFFF" if severity in ("Blocker", "Critical") else "000000")
                    cell.alignment = center

                # Status 컬럼
                elif col_idx == 12:
                    cell.fill = status_fill
                    cell.alignment = center

                # TC_ID 컬럼
                elif col_idx == 1:
                    cell.font = Font(name="Arial", bold=True, size=9)
                    cell.alignment = center

                else:
                    cell.font = Font(name="Arial", size=9)

            ws.row_dimensions[row_idx].height = 45

        # 헤더 고정
        ws.freeze_panes = "A2"

        # 요약 시트 추가
        ws_summary = wb.create_sheet("Summary")
        self._create_summary_sheet(ws_summary, tc_list, header_font, header_fill, border, center)

        wb.save(output_path)
        return output_path

    def _create_summary_sheet(self, ws, tc_list, header_font, header_fill, border, center):
        """Summary 시트 생성"""
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 15

        # 타이틀
        ws["A1"] = "QA-Flow TC 요약"
        ws["A1"].font = Font(name="Arial", bold=True, size=13, color="1F4E79")
        ws["A1"].alignment = Alignment(horizontal="left")
        ws["A2"] = f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A2"].font = Font(name="Arial", size=9, color="888888")

        # 통계
        stats_start = 4
        headers = ["항목", "건수"]
        for i, h in enumerate(headers, start=1):
            cell = ws.cell(row=stats_start, column=i, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        total = len(tc_list)
        priority_counts = {}
        for tc in tc_list:
            p = tc.get("priority", "Major")
            priority_counts[p] = priority_counts.get(p, 0) + 1

        rows = [("전체 TC", total)] + [(f"Priority: {k}", v) for k, v in priority_counts.items()]
        for i, (label, val) in enumerate(rows, start=stats_start + 1):
            ws.cell(row=i, column=1, value=label).border = border
            ws.cell(row=i, column=2, value=val).border = border
            ws.cell(row=i, column=2).alignment = center


def upload_to_jira(tc_list: list, output_path: str = "") -> dict:
    """생성된 TC를 Jira 이슈로 등록"""
    from dotenv import load_dotenv
    from pathlib import Path
    import requests
    from requests.auth import HTTPBasicAuth

    load_dotenv(dotenv_path=Path.cwd() / ".env", encoding="utf-8")

    jira_url = os.getenv("JIRA_URL", "").rstrip("/")
    jira_email = os.getenv("JIRA_EMAIL", "")
    jira_token = os.getenv("JIRA_API_TOKEN", "")
    jira_project = os.getenv("JIRA_PROJECT_KEY", "")

    if not all([jira_url, jira_email, jira_token, jira_project]):
        console.print("[red]Jira 설정이 없습니다. qa-flow init을 다시 실행하거나 .env를 확인하세요.[/red]")
        return {}

    auth = HTTPBasicAuth(jira_email, jira_token)
    headers = {"Accept": "application/json", "Content-Type": "application/json"}

    # Jira Priority 매핑
    priority_map = {
        "Blocker":  "Highest",
        "Critical": "High",
        "Major":    "Medium",
        "Minor":    "Low",
        "Low":      "Lowest",
    }

    created_issues = {}
    failed = []

    with console.status("[bold green]Jira에 이슈를 등록하고 있습니다..."):
        for i, tc in enumerate(tc_list):
            tc_id = f"TC-{i + 1:03d}"
            priority_name = priority_map.get(tc.get("priority", "Major"), "Medium")

            description = (
                f"*모듈*: {tc.get('module', '')}\n\n"
                f"*사전 조건*:\n{tc.get('precondition', '')}\n\n"
                f"*테스트 단계*:\n{tc.get('test_steps', '')}\n\n"
                f"*테스트 데이터*:\n{tc.get('test_data', '')}\n\n"
                f"*예상 결과*:\n{tc.get('expected_result', '')}\n\n"
                f"*Severity*: {tc.get('severity', '')}\n"
                f"*Test Type*: {tc.get('test_type', '')}\n"
                f"*TC ID*: {tc_id}"
            )

            payload = {
                "fields": {
                    "project": {"key": jira_project},
                    "summary": f"[{tc_id}] {tc.get('test_scenario', '')}",
                    "description": description,
                    "issuetype": {"name": "Test"},
                    "priority": {"name": priority_name},
                }
            }

            try:
                res = requests.post(
                    f"{jira_url}/rest/api/2/issue",
                    json=payload,
                    auth=auth,
                    headers=headers,
                    timeout=10
                )
                if res.status_code == 201:
                    issue_key = res.json().get("key", "")
                    created_issues[tc_id] = issue_key
                else:
                    failed.append(tc_id)
            except Exception:
                failed.append(tc_id)

    return {"created": created_issues, "failed": failed}
